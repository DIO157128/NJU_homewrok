import csv
from openpyxl import Workbook


def transfer_polarity_to_bit(polarity):
    if polarity == "positive":
        return 1
    if polarity == "neutral":
        return 0
    if polarity == "negative":
        return -1


def dataload(file_path):
    workbook = Workbook()
    sheet = workbook.active

    text_list = []
    label_list = []

    # 获取filename
    filename = file_path.split("\\")[-1].split(".")[0] + ".xlsx"

    with open(file_path, "r", encoding="utf-8") as dataset:
        data_reader = csv.reader(dataset)

        for line in data_reader:
            temp = str(line).split(";")
            label_list.append(int(transfer_polarity_to_bit(temp[1])))
            if temp[2].endswith("]"):
                temp[2] = temp[2][:len(temp[2]) - 1]
            text_list.append(str(temp[2]))

    for index, value in enumerate(text_list, start=1):
        sheet.cell(row=index, column=1, value=value)

    for index, value in enumerate(label_list, start=1):
        sheet.cell(row=index, column=2, value=value)

    workbook.save(filename)


if __name__ == "__main__":

    # 请使用绝对地址
    train_set = ""
    test_set = ""

    dataload(train_set)
    dataload(test_set)