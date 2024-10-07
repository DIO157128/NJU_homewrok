import argparse
import csv
import logging
import os
import re

import nltk
import numpy as np
from cleanlab.classification import CleanLearning
from cleanlab.filter import find_label_issues
from imblearn.over_sampling import SVMSMOTE
from nltk import SnowballStemmer
from openpyxl import load_workbook
from sklearn.ensemble import VotingClassifier, GradientBoostingClassifier, RandomForestClassifier
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import SGDClassifier
from sklearn.metrics import precision_score, recall_score, f1_score, accuracy_score, confusion_matrix
from sklearn.model_selection import cross_val_predict

import sentiCR


class Senti_24(sentiCR.SentiCR):

        # 加载训练集训练模型
        def create_model_from_training_data(self):
            training_comments = []
            training_ratings = []
            print("Training classifier model..")
            for sentidata in self.training_data:
                comments = sentiCR.preprocess_text(sentidata.text)
                training_comments.append(comments)
                training_ratings.append(sentidata.rating)

            # 提取特征
            # discard stopwords, apply stemming, and discard words present in less than 3 comments
            # max_df 丢弃在超过50%的评论中出现的词语，将其视为常见词
            # min_df 丢弃在不超过3个评论中出现的词语，将其视为低频词
            self.vectorizer = TfidfVectorizer(tokenizer=sentiCR.tokenize_and_stem, sublinear_tf=True, max_df=0.5,
                                              stop_words=mystop_words, min_df=3)
            X_train = self.vectorizer.fit_transform(training_comments).toarray()
            Y_train = np.array(training_ratings)

            # 改变少数类的比例
            # Apply SMOTE to improve ratio of the minority class
            # smote_model = SMOTE(sampling_strategy=0.5, random_state=None,  k_neighbors=15, n_jobs=1)

            smote_model = SVMSMOTE(random_state=None, k_neighbors=15, svm_estimator=None)

            X_resampled, Y_resampled = smote_model.fit_resample(X_train, Y_train)
            Y_resampled_ecoder = []
            for i in Y_resampled:
                Y_resampled_ecoder += [int(i)]

            model = self.get_classifier()
            model.fit(X_resampled, Y_resampled_ecoder)
            # print(X_resampled)
            # print(Y_resampled)
            # predProbs = cross_val_predict(
            #     model,
            #     X_resampled,
            #     Y_resampled,
            #     cv=5,
            #     method="predict_proba",
            # )
            # y_data_int = []
            # for i in Y_resampled:
            #     y_data_int += [int(i)+1]
            # ranked_label_issues = find_label_issues(
            #     y_data_int,
            #     predProbs,
            #     return_indices_ranked_by="self_confidence",
            # )
            # print(len(ranked_label_issues))
            # clean_X = np.delete(X_resampled, list(ranked_label_issues), 0)
            # clean_y = np.delete(Y_resampled, list(ranked_label_issues), 0)
            # print(clean_X.shape, clean_y.shape)
            #
            # # 在处理之后的数据再弄一次决策树
            # model2 = VotingClassifier(
            #     estimators=[('GBT', GradientBoostingClassifier()), ('sgd', SGDClassifier()),
            #                 ('rf', RandomForestClassifier())],
            #     voting='hard')
            # model2.fit(clean_X, clean_y)
            # print('finish retraining')

            return model

# 词干提取
stemmer = SnowballStemmer("english")

mystop_words = [
    'i', 'me', 'my', 'myself', 'we', 'our', 'ourselves', 'you', 'your',
    'yourself', 'yourselves', 'he', 'him', 'his', 'himself', 'she', 'her',
    'herself', 'it', 'its', 'itself', 'they', 'them', 'their', 'themselves',
    'this', 'that', 'these', 'those', 'am', 'is', 'are', 'was', 'were', 'be', 'been', 'being',
    'have', 'has', 'had', 'having', 'do', 'does', 'did', 'doing', 'a', 'an', 'the',
    'and', 'if', 'or', 'as', 'until', 'of', 'at', 'by', 'between', 'into',
    'through', 'during', 'to', 'from', 'in', 'out', 'on', 'off', 'then', 'once', 'here',
    'there', 'all', 'any', 'both', 'each', 'few', 'more',
    'other', 'some', 'such', 'than', 'too', 'very', 's', 't', 'can', 'will', 'don', 'should', 'now'
    # keywords
                                                                                              'while', 'case', 'switch',
    'def', 'abstract', 'byte', 'continue', 'native', 'private', 'synchronized',
    'if', 'do', 'include', 'each', 'than', 'finally', 'class', 'double', 'float', 'int', 'else', 'instanceof',
    'long', 'super', 'import', 'short', 'default', 'catch', 'try', 'new', 'final', 'extends', 'implements',
    'public', 'protected', 'static', 'this', 'return', 'char', 'const', 'break', 'boolean', 'bool', 'package',
    'byte', 'assert', 'raise', 'global', 'with', 'or', 'yield', 'in', 'out', 'except', 'and', 'enum', 'signed',
    'void', 'virtual', 'union', 'goto', 'var', 'function', 'require', 'print', 'echo', 'foreach', 'elseif', 'namespace',
    'delegate', 'event', 'override', 'struct', 'readonly', 'explicit', 'interface', 'get', 'set', 'elif', 'for',
    'throw', 'throws', 'lambda', 'endfor', 'endforeach', 'endif', 'endwhile', 'clone'
]


emodict = []
contractions_dict = []


# Read in the words with sentiment from the dictionary
with open("Contractions.txt", "r") as contractions, \
        open("EmoticonLookupTable.txt", "r") as emotable:
    contractions_reader = csv.reader(contractions, delimiter='\t')
    emoticon_reader = csv.reader(emotable, delimiter='\t')

    # Hash words from dictionary with their values
    contractions_dict = {rows[0]: rows[1] for rows in contractions_reader}
    emodict = {rows[0]: rows[1] for rows in emoticon_reader}

    contractions.close()
    emotable.close()

grammar = r"""
NegP: {<VERB>?<ADV>+<VERB|ADJ>?<PRT|ADV><VERB>}
{<VERB>?<ADV>+<VERB|ADJ>*<ADP|DET>?<ADJ>?<NOUN>?<ADV>?}

"""
chunk_parser = nltk.RegexpParser(grammar)

# 生成缩略词替换正则表达式
contractions_regex = re.compile('(%s)' % '|'.join(contractions_dict.keys()))


# url匹配
url_regex = re.compile('http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+')


negation_words = ['not', 'never', 'none', 'nobody', 'nowhere', 'neither', 'barely', 'hardly',
                  'nothing', 'rarely', 'seldom', 'despite']

emoticon_words = ['PositiveSentiment', 'NegativeSentiment']


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Supervised sentiment classifier')

    parser.add_argument('--algo', type=str,
                        help='Classification algorithm', default="MIX")

    logging.config.fileConfig("../config/logging.conf")
    logger = logging.getLogger()
    os.makedirs(f"logs/improvement", exist_ok=True)
    save_path = "improvement/MIX_0_clean"
    fileHandler = logging.FileHandler("logs/{}.log".format(save_path), mode="w")
    formatter = logging.Formatter(
        "[%(levelname)s] %(filename)s:%(lineno)d > %(message)s"
    )
    fileHandler.setFormatter(formatter)
    logger.addHandler(fileHandler)

    args = parser.parse_args()
    ALGO = args.algo

    print("Cross validation")
    print("Algrithm: " + ALGO)

    workbook_train = load_workbook("./train3098itemPOLARITY.xlsx")
    sheet_train = workbook_train.worksheets[0]
    workbook_test = load_workbook("./test1326itemPOLARITY.xlsx")
    sheet_test = workbook_test.worksheets[0]
    oracle_data = []
    test_data = []


    for row in sheet_train.iter_rows(min_row=1, max_row=sheet_train.max_row, values_only=True):
        comments = sentiCR.SentimentData(row[0], row[1])
        oracle_data.append(comments)

    for row in sheet_test.iter_rows(min_row=1, max_row=sheet_test.max_row, values_only=True):
        comments = sentiCR.SentimentData(row[0], row[1])
        test_data.append(comments)


    oracle_data = np.array(oracle_data)
    test_data = np.array(test_data)

    # 生成模型
    classifier_model = Senti_24(algo=ALGO, training_data=oracle_data)

    # 处理测试集
    test_comments = [comments.text for comments in test_data]
    test_ratings = [comments.rating for comments in test_data]
    test_ratings_encoder = []
    for i in test_ratings:
        test_ratings_encoder += [int(i)]

    pred = classifier_model.get_sentiment_polarity_collection(test_comments)

    class_labels = ['positive', 'neutral', 'negative']
    conf_matrix = confusion_matrix(test_ratings_encoder, pred)
    print(conf_matrix)

    for i in range(3):
        TP = conf_matrix[i, i]
        FN = conf_matrix[i, :].sum() - TP
        FP = conf_matrix[:, i].sum() - TP

        recall = TP / (TP + FN)
        precision = TP / (TP + FP)
        f1score = 2 * (precision * recall) / (precision + recall)
        logger.info("-------------------------")
        logger.info("recall_of_" + class_labels[i] + ": " + str(recall))
        logger.info("precision_of_" + class_labels[i] + ": " + str(precision))
        logger.info("f1score_of_" + class_labels[i] + ": " + str(f1score))
        logger.info("-------------------------")

    accuracy = accuracy_score(test_ratings_encoder, pred)
    logger.info("Accuracy:" + str(accuracy))








