import csv
import os
import re
import numpy as np
import pandas as pd
from openpyxl import load_workbook


def removeAt(path):
    name = path.split("/")[1].split(".")[0]
    # 读取原始Excel文件
    workbook = load_workbook(path)
    sheet = workbook.active

    pattern = r'@[\w\-.]+'

    # 遍历第一列的每个单元格，去除文本中的用户名
    for cell in sheet['A']:
        if cell.value:
            cell.value = re.sub(pattern, '', cell.value)
    folder_path = "withoutAt"
    if not os.path.exists(folder_path):
        # 创建文件夹
        os.makedirs(folder_path)
    # 保存为新的Excel文件
    workbook.save('./{}/wo@_{}.xlsx'.format(folder_path, name))


class SplitObject:
    def __init__(self, half1, half2):
        self.half1 = half1
        self.half2 = half2


def pairEmos(excel_path,name):
    # 加载情感词典
    emotion_table_path = 'sentistrength_jar/SentStrength_Data/EmotionLookupTable.txt'
    emotion_table = open(emotion_table_path, 'r')
    emotion_reader = csv.reader(emotion_table, delimiter='\t')
    emotion_dict = {rows[0]: int(int(rows[1]) / abs(int(rows[1]))) for rows in emotion_reader}
    emotion_dict['PositiveSentiment'] = 1
    emotion_dict['NegativeSentiment'] = -1
    # 加载表情符号词典
    emoticon_table_path = 'EmoticonLookupTable.txt'
    emoticon_table = open(emoticon_table_path, 'r')
    emoticon_reader = csv.reader(emoticon_table, delimiter='\t')
    emoticon_dict = {rows[0]: rows[1] for rows in emoticon_reader}
    #打开excel
    workbook = load_workbook(excel_path)
    sheet = workbook.active
    text = []
    label = []
    postive_list = []
    negative_list = []

    # 读取text和label
    for cell in sheet[chr(64 + 1)]:
        text.append(cell.value)
    for cell in sheet[chr(64 + 2)]:
        label.append(cell.value)
    print(len(text))
    for i in range(len(text)):
        t = text[i]
        # 去除表情符号
        for j, k in emoticon_dict.items():
            t = t.replace(j, k)
        split_t = t.split()
        l = label[i]
        # 是积极的就加到积极的list里
        if l == 1:
            tem_postive = []
            for o in range(len(split_t)):
                single_word = split_t[o]
                if single_word in emotion_dict.keys() and emotion_dict[single_word] == 1:
                    tem_postive.append(o)
            if len(tem_postive)!=0 and len(tem_postive) % 2 == 0:
                mid_index = int(
                    (tem_postive[int(len(tem_postive) / 2)] + tem_postive[int(len(tem_postive) / 2 - 1)]) / 2)
                postive_list.append(SplitObject(" ".join(split_t[:mid_index]), " ".join(split_t[mid_index:])))
        elif l ==-1:
            tem_negative = []
            for o in range(len(split_t)):
                single_word = split_t[o]
                if single_word in emotion_dict.keys() and emotion_dict[single_word] == -1:
                    tem_negative.append(o)
            if len(tem_negative)!=0 and len(tem_negative) % 2 == 0:
                mid_index = int(
                    (tem_negative[int(len(tem_negative) / 2)] + tem_negative[int(len(tem_negative) / 2 - 1)]) / 2)
                negative_list.append(SplitObject(" ".join(split_t[:mid_index]), " ".join(split_t[mid_index:])))
    for i in range(len(postive_list)):
        for j in range(len(postive_list)):
            if i!=j:
                p1 = postive_list[i]
                p2 = postive_list[j]
                text.append(p1.half1+" "+p2.half2)
                label.append(1)
    for i in range(len(negative_list)):
        for j in range(len(negative_list)):
            if i!=j:
                p1 = negative_list[i]
                p2 = negative_list[j]
                text.append(p1.half1+" "+p2.half2)
                label.append(-1)
    print(len(text))
    df = pd.DataFrame({'col1': text, 'col2': label})
    folder_path = "splitAugmentation"
    if not os.path.exists(folder_path):
        # 创建文件夹
        os.makedirs(folder_path)
    # 将DataFrame写入Excel文件，不包含列名
    df_shuffled = df.sample(frac=1, random_state=42)

    # 重置索引
    df_shuffled.reset_index(drop=True, inplace=True)
    df_shuffled.to_excel(folder_path+"/{}.xlsx".format(name), index=False, header=False)

if __name__ == '__main__':
    removeAt("original_data/train3098itemPOLARITY.xlsx")
    pairEmos("withoutAt/wo@_train3098itemPOLARITY.xlsx", 'train3098itemPOLARITY')
