# coding utf-8
import os

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.metrics import confusion_matrix, accuracy_score


def run_score(path, name):
    workbook = load_workbook(path)
    sheet = workbook.worksheets[0]
    source = []
    target = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
        source.append(row[0])
        target.append(row[1])
    f = open("{}.txt".format(name), "a", encoding="utf-8")
    for t in source:
        f.write(t + "\n")
    f.close()
    os.system(
        "java -jar ./sentistrength_jar/sentistrength.jar sentidata ./sentistrength_jar/SentStrength_Data/ input {}.txt trinary".format(
            name))
    # 手动给结果添加# -*- coding: utf-8 -*-
    f = open("{}0_out.txt".format(name), 'r', encoding="utf-8")
    score = []
    lines = f.readlines()
    for l in lines[2:]:
        score.append(l.split()[0])
    score = np.array(score)
    df = pd.DataFrame()
    df['text'] = source
    df['polarity'] = target
    df['Sentistrength score'] = score
    df.to_csv("{}_sentistrength.csv".format(name))


def calculate(path):
    df = pd.read_csv(path)
    polarity = df['polarity']
    score = df['Sentistrength score']
    result = np.array(score)
    target = np.array(polarity)
    class_labels = [-1,0,1]
    conf_matrix = confusion_matrix(target, result)
    print(conf_matrix)
    for i in range(3):
        print("Class:"+str(class_labels[i]))
        TP = conf_matrix[i, i]
        FN = conf_matrix[i, :].sum() - TP
        FP = conf_matrix[:, i].sum() - TP

        recall = TP / (TP + FN)
        precision = TP / (TP + FP)
        f1score = 2 * (precision * recall) / (precision + recall)
        print("Recall:"+str(recall))
        print("Precision:"+str(precision))
        print("F1-score:"+str(f1score))
        print()

    accuracy = accuracy_score(target, result)
    print("Accuracy:"+str(accuracy))


if __name__ == '__main__':
    # run_score("./data/train3098itemPOLARITY.xlsx", "train")
    # run_score("./data/test1326itemPOLARITY.xlsx", "test")
    # calculate("train_sentistrength.csv")
    # print("*"*200)
    calculate("sentistrength_result/test_sentistrength.csv")
