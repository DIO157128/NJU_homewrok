## 根据文件的路径加载数据集
import numpy as np
from openpyxl import load_workbook
import preprocess.process_text as preprocess


def load_data(train_path, test_path):
    workbook_train = load_workbook(train_path)
    sheet_train = workbook_train.worksheets[0]
    workbook_test = load_workbook(test_path)
    sheet_test = workbook_test.worksheets[0]
    X_train = []
    y_train = []

    X_test = []
    y_test = []

    for row in sheet_test.iter_rows(min_row=1, max_row=sheet_test.max_row, values_only=True):
        if row[1] == -1:
            y_test.append(2)
        else :
            y_test.append(row[1])

        X_test.append(preprocess.preprocess_text(row[0]))

    for row in sheet_train.iter_rows(min_row=1, max_row=sheet_train.max_row, values_only=True):
        if row[1] == -1:
            y_train.append(2)
        else:
            y_train.append(row[1])

        X_train.append(preprocess.preprocess_text(row[0]))

    return np.array(X_train), np.array(X_test), np.array(y_train), np.array(y_test)

